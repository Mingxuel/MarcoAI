"""
股票池（STOCK_CODES_ALL）管理模块

数据来源:
    MarcoAI/AIData/INFO/SZ100.xlsx —— 股票池明细（含股票代码、名称、流通股本）

存储文件:
    MarcoAI/AIData/STOCK_CODES_ALL —— 每行一条: 代码|名称|流通股本(股)
    示例: 000001.SZ|平安银行|19405600768

API 说明:
    STOCK_CODES() -> list[str]
        返回全部股票代码列表（带 .SZ/.SH 后缀），带缓存，文件缺失时自动更新
        示例: ['000001.SZ', '600000.SH', ...]

    UPDATE_STOCK_CODES() -> list[str]
        从 SZ100.xlsx 重新提取股票池，清空 STOCK_CODES_ALL 文件后写入，
        并同步刷新缓存；返回完整的 "代码|名称|流通股本" 行列表

    STOCK_CODES_ALL() -> list[str]
        返回完整股票池行列表（"代码|名称|流通股本"），从 STOCK_CODES_ALL 文件读取，
        文件缺失时回退 UPDATE_STOCK_CODES() 生成
        示例: ['000001.SZ|平安银行|19405600768', ...]

    GET_STOCK_INFO(stock_code: str) -> tuple[str, int] | None
        通过股票代码获取 (股票名称, 流通股本)，支持带/不带后缀（6 开头自动补 .SH），
        查不到返回 None
        示例: GET_STOCK_INFO("000001") -> ('平安银行', 19405600768)
"""

import os
import sys

_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
if _root not in sys.path:
    sys.path.insert(0, _root)
from AICode.MarcoAPI.Update.Constants import *
from AICode.MarcoAPI.Update.Path import *

from openpyxl import load_workbook

_STOCK_CODES_CACHE: dict[str, list[str]] = {}
_STOCK_INFO_CACHE: dict[str, dict[str, tuple[str, int]]] = {}

def STOCK_CODES():
    if "data" not in _STOCK_CODES_CACHE:
        if os.path.exists(PATH_AIDATA_STOCK_CODES_ALL()):
            with open(PATH_AIDATA_STOCK_CODES_ALL(), "r", encoding="utf-8") as file:
                _STOCK_CODES_CACHE["data"] = [line.split("|")[0] for line in file.read().splitlines() if line]
        else:
            _STOCK_CODES_CACHE["data"] = [line.split("|")[0] for line in UPDATE_STOCK_CODES()]
    return _STOCK_CODES_CACHE["data"]

def UPDATE_STOCK_CODES():
    """从 SZ100.xlsx 股票池提取股票代码/名称/流通股本，清空 STOCK_CODES_ALL 文件后写入"""
    data: list[str] = []
    workbook = load_workbook(PATH_AIDATA_INFO_SZ100())
    try:
        sheet = workbook.active
        if sheet is None:
            return data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 10 or row[1] is None:
                continue
            code = str(row[1]).strip()
            if not code.isdigit() or len(code) != 6:
                continue
            name = str(row[2]).strip() if row[2] is not None else ""
            float_share = int(row[9]) if isinstance(row[9], (int, float)) else 0
            code = f"{code}.SH" if code.startswith("6") else f"{code}.SZ"
            data.append(f"{code}|{name}|{float_share}")
    finally:
        workbook.close()
    with open(PATH_AIDATA_STOCK_CODES_ALL(), "w", encoding="utf-8") as file:
        file.write("\n".join(data))
    _STOCK_CODES_CACHE["data"] = [line.split("|")[0] for line in data]
    return data

def STOCK_CODES_ALL() -> list[str]:
    """返回完整股票池行列表（"代码|名称|流通股本"），从 STOCK_CODES_ALL 文件读取，缺失时回退生成"""
    if os.path.exists(PATH_AIDATA_STOCK_CODES_ALL()):
        with open(PATH_AIDATA_STOCK_CODES_ALL(), "r", encoding="utf-8") as file:
            return [line for line in file.read().splitlines() if line]
    return UPDATE_STOCK_CODES()

def GET_STOCK_INFO(stock_code: str) -> tuple[str, int] | None:
    """通过股票代码获取 (股票名称, 流通股本)，支持带 .SZ/.SH 后缀或不带后缀，找不到返回 None"""
    code = stock_code.strip().upper()
    if "." not in code and len(code) == 6 and code.isdigit():
        code = f"{code}.SH" if code.startswith("6") else f"{code}.SZ"
    if "data" not in _STOCK_INFO_CACHE:
        if os.path.exists(PATH_AIDATA_STOCK_CODES_ALL()):
            with open(PATH_AIDATA_STOCK_CODES_ALL(), "r", encoding="utf-8") as file:
                _STOCK_INFO_CACHE["data"] = _LOAD_STOCK_INFO(file.read())
        else:
            _STOCK_INFO_CACHE["data"] = _LOAD_STOCK_INFO("\n".join(UPDATE_STOCK_CODES()))
    return _STOCK_INFO_CACHE["data"].get(code)

def _LOAD_STOCK_INFO(content: str) -> dict[str, tuple[str, int]]:
    """解析 STOCK_CODES_ALL 内容为 {代码: (名称, 流通股本)}"""
    info_map: dict[str, tuple[str, int]] = {}
    for line in content.splitlines():
        if not line:
            continue
        parts = line.split("|")
        if len(parts) >= 3:
            info_map[parts[0]] = (parts[1], int(parts[2]))
    return info_map

if __name__ == "__main__":
    print(UPDATE_STOCK_CODES())
