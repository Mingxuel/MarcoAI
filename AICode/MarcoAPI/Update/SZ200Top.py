"""
涨停股票（TOP）计算与管理模块

数据来源:
    MarcoAI/AIData/TOP_ORIGIN —— 每日涨停股原始 xlsx（文件名含日期）

存储文件:
    MarcoAI/AIData/TOP     —— 每日涨停股列表，文件名为交易日（YYYYMMDD），每行一个代码

API 说明:
    UPDATE_TOP() -> None
        遍历 TOP_ORIGIN 下的涨停 xlsx，提取股票代码（6 开头加 .SH，其余加 .SZ），
        以涨停日期命名写入 TOP 目录；多进程并行处理

    GET_TOP() -> dict[str, set[str]]
        返回跨进程共享的 Manager.dict 代理（{交易日: 涨停代码集合}），供 worker 进程使用

    IS_TOP(stock_code, trading_date) -> bool
        判断某股票在指定交易日是否涨停；优先读缓存（主进程/worker），兜底读文件

    GENERATE_TOP
        UPDATE_TOP 的 worker 函数，供 ProcessPoolExecutor 并行调度
"""

import os
import re
import sys
import warnings
from concurrent.futures import ProcessPoolExecutor, as_completed
from multiprocessing import Manager
from typing import Optional
from .progress import ProgressBar as tqdm

from openpyxl import load_workbook

_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
if _root not in sys.path:
    sys.path.insert(0, _root)
from AICode.MarcoAPI.Update.Constants import *
from AICode.MarcoAPI.Update.StockCodes import *
from AICode.MarcoAPI.Update.TradingDates import *
from AICode.MarcoAPI.Update.Path import *

_TOP_CACHE: dict[str, set[str]] = {}
_WORKER_TOP_CACHE: Optional[dict[str, set[str]]] = None  # worker 进程中由 initializer 注入

def INIT_TOP(proxy: dict[str, set[str]]):
    """ProcessPoolExecutor initializer: 将 Manager.dict 代理注入 worker 全局变量"""
    global _WORKER_TOP_CACHE
    _WORKER_TOP_CACHE = proxy  # pyright: ignore[reportConstantRedefinition]

_MANAGER: Optional[Manager] = None

def GET_TOP() -> dict[str, set[str]]:
    """创建并返回跨进程共享的 Manager.dict 代理（从已填充的 _TOP_CACHE 构建）"""
    global _MANAGER
    if _MANAGER is None:
        _MANAGER = Manager()  # pyright: ignore[reportConstantRedefinition]
    proxy: dict[str, set[str]] = _MANAGER.dict()  # pyright: ignore[reportOptionalMemberAccess]
    proxy.update(_TOP_CACHE)
    return proxy

def UPDATE_TOP():
    """从 TOP_ORIGIN 的每日涨停 xlsx 生成 TOP 目录下的涨停股列表文件（代码带 .SZ/.SH 后缀）"""
    top_origin_dir = PATH_AIDATA_TOP_ORIGIN()
    if not os.path.exists(top_origin_dir):
        print(f"UPDATE_TOP: TOP_ORIGIN 目录不存在 {top_origin_dir}")
        return
    xlsx_files = [
        os.path.join(top_origin_dir, file_name)
        for file_name in sorted(os.listdir(top_origin_dir))
        if file_name.endswith(".xlsx") and re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", file_name) is not None
    ]
    with ProcessPoolExecutor(max_workers=32) as pool:
        future_to_file = {pool.submit(GENERATE_TOP, f): f for f in xlsx_files}
        with tqdm(total=len(future_to_file), desc="TOP", ncols=90, disable=False) as bar:
            for fut in as_completed(future_to_file):
                fut.result()
                bar.set_postfix(file=os.path.basename(future_to_file[fut]), refresh=False)
                bar.update(1)
    return f"TOP: 完成 {len(xlsx_files)} 个涨停文件"

def GENERATE_TOP(xlsx_file: str):
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", os.path.basename(xlsx_file))
    if match is None:
        return
    trading_date = f"{match.group(1)}{int(match.group(2)):02d}{int(match.group(3)):02d}"
    top_file = f"{PATH_AIDATA_TOP()}/{trading_date}"
    stock_codes = _READ_TOP_ORIGIN(xlsx_file)
    with open(top_file, "w") as file:
        for stock_code in stock_codes:
            file.write(f"{stock_code}\n")

def _READ_TOP_ORIGIN(xlsx_file: str) -> list[str]:
    """读取涨停 xlsx，返回带后缀的股票代码列表（去重、保持顺序）"""
    stock_codes: list[str] = []
    seen: set[str] = set()
    with warnings.catch_warnings():
        # 抑制 openpyxl 读取部分 xlsx 时的无害警告：
        # "Workbook contains no default style, apply openpyxl's default"
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(xlsx_file)
    try:
        sheet = workbook.active
        if sheet is None:
            return stock_codes
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 2 or row[1] is None:
                continue
            code = str(row[1]).strip()
            if code.endswith(".0"):
                code = code[:-2]
            if not code.isdigit() or len(code) != 6:
                continue
            code = f"{code}.SH" if code.startswith("6") else f"{code}.SZ"
            if code not in seen:
                seen.add(code)
                stock_codes.append(code)
    finally:
        workbook.close()
    return stock_codes

def IS_TOP(stock_code: str, trading_date: str):
    # 主进程：使用模块级普通 dict
    stock_set = _TOP_CACHE.get(trading_date)
    if stock_set is not None:
        return stock_code in stock_set
    # worker 进程：使用 Manager.dict 代理（通过 initializer 注入）
    if _WORKER_TOP_CACHE is not None:
        stock_set = _WORKER_TOP_CACHE.get(trading_date)
        if stock_set is not None:
            return stock_code in stock_set
    # 缓存未填充时，回退到文件读取（兜底）
    top_file = f"{PATH_AIDATA_TOP()}/{trading_date}"
    if not os.path.exists(top_file):
        return False
    with open(top_file, "r") as file:
        for line in file:
            if line.strip() == stock_code:
                return True
    return False

if __name__ == "__main__":
    UPDATE_TOP()
